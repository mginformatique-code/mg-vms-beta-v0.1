"""MG-VMS — Module Rapports (CSV / Excel / PDF).

Génère des rapports filtrés (plage de dates + site, cloisonnés) pour :
plaques (ANPR), événements, alertes, équipements réseau.
"""
import io
import csv
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from database import db
from auth import get_current_user, require_role, log_audit, site_scope, allowed_sites

reports_router = APIRouter(prefix="/api/reports", tags=["reports"])

# type -> (label, collection, colonnes [(clé, en-tête)], champ_date)
REPORTS = {
    "plates": ("Rapport ANPR / Plaques", "plates", [
        ("plate", "Plaque"), ("timestamp", "Date"), ("camera_name", "Caméra"), ("site_name", "Site"),
        ("vehicle_color", "Couleur"), ("vehicle_make", "Marque"), ("vehicle_model", "Modèle"),
        ("vehicle_type", "Type"), ("direction", "Sens"), ("confidence", "Confiance"), ("list_status", "Liste"),
    ], "timestamp"),
    "events": ("Rapport Événements IA", "events", [
        ("type", "Type"), ("timestamp", "Date"), ("camera_name", "Caméra"),
        ("site_name", "Site"), ("confidence", "Confiance"),
    ], "timestamp"),
    "alerts": ("Rapport Alertes", "alerts", [
        ("severity", "Sévérité"), ("type", "Type"), ("message", "Message"), ("camera_name", "Caméra"),
        ("site_name", "Site"), ("acknowledged", "Acquittée"), ("timestamp", "Date"),
    ], "timestamp"),
    "equipment": ("Rapport Supervision réseau", "equipment", [
        ("name", "Nom"), ("type", "Type"), ("ip", "IP"), ("site_name", "Site"),
        ("status", "Statut"), ("latency_ms", "Latence (ms)"), ("vendor", "Fabricant"), ("model", "Modèle"),
    ], None),
}


@reports_router.get("/types")
async def report_types(user: dict = Depends(require_role("technician"))):
    return [{"key": k, "label": v[0], "date_filter": v[3] is not None} for k, v in REPORTS.items()]


async def _fetch_rows(report_type: str, user: dict, site_id, date_from, date_to):
    label, coll, cols, date_field = REPORTS[report_type]
    q = {}
    if site_id:
        q["site_id"] = site_id
    site_scope(q, user)
    if date_field and (date_from or date_to):
        rng = {}
        if date_from:
            rng["$gte"] = date_from
        if date_to:
            rng["$lte"] = date_to
        q[date_field] = rng
    sort_field = date_field or "name"
    rows = await getattr(db, coll).find(q, {"_id": 0}).sort(sort_field, -1).to_list(5000)
    return label, cols, rows


def _cell(row, key):
    v = row.get(key, "")
    if isinstance(v, bool):
        return "Oui" if v else "Non"
    return "" if v is None else str(v)


@reports_router.get("/{report_type}")
async def generate_report(report_type: str, format: str = Query("csv"),
                          site_id: Optional[str] = None, date_from: Optional[str] = None,
                          date_to: Optional[str] = None, user: dict = Depends(require_role("technician"))):
    if report_type not in REPORTS:
        raise HTTPException(404, "Type de rapport inconnu")
    if format not in ("csv", "xlsx", "pdf"):
        raise HTTPException(400, "Format invalide")
    label, cols, rows = await _fetch_rows(report_type, user, site_id, date_from, date_to)
    headers = [c[1] for c in cols]
    keys = [c[0] for c in cols]
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    fname = f"mgvms_{report_type}_{stamp}.{format}"
    await log_audit(user, "report_generated", report_type, f"{format} · {len(rows)} lignes")

    if format == "csv":
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(headers)
        for r in rows:
            w.writerow([_cell(r, k) for k in keys])
        out.seek(0)
        return StreamingResponse(iter([out.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": f"attachment; filename={fname}"})

    if format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = report_type[:31]
        ws.append(headers)
        hdr_fill = PatternFill("solid", fgColor="0044FF")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
        for r in rows:
            ws.append([_cell(r, k) for k in keys])
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, min(40, len(h) + 6))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(iter([buf.getvalue()]),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={fname}"})

    # pdf
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.2 * cm, bottomMargin=1 * cm,
                            leftMargin=1 * cm, rightMargin=1 * cm)
    styles = getSampleStyleSheet()
    elems = [Paragraph(f"MG-VMS — {label}", styles["Title"])]
    sub = f"Généré le {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')} · {len(rows)} enregistrement(s)"
    if date_from or date_to:
        sub += f" · Période : {date_from or '—'} → {date_to or '—'}"
    elems += [Paragraph(sub, styles["Normal"]), Spacer(1, 10)]
    data = [headers] + [[(_cell(r, k)[:40]) for k in keys] for r in rows[:1500]]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0044FF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f5ff")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems.append(table)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})
