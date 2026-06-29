"""Sprint Itération 9 — Tests Rapports + ANPR blacklist enrichie + Poll réseau périodique.

Couvre:
- /api/reports/types et /api/reports/{type}?format=csv|xlsx|pdf
- Filtres date_from/date_to/site_id, et 'equipment' sans filtre date
- Sécurité: viewer (readonly) -> 403 (require_role technician)
- POST /api/anpr/detect ZZ-999-ZZ -> alerte critique sans 500
- Poll réseau périodique (broadcaster) -> /api/network/stats reste cohérent
- Régressions: /api/network/poll + /api/recordings/export
"""
import io
import os
import time
import csv as _csv
import zipfile
import pytest
import requests
from openpyxl import load_workbook

# Lecture BASE_URL depuis env ou frontend/.env
BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")

API = f"{BASE_URL}/api"

ADMIN = ("admin@mg-vms.com", "Admin@2026")
TECH = ("tech@mg-vms.com", "Tech@2026")
VIEWER = ("viewer@mg-vms.com", "Viewer@2026")


def _login(email, pwd):
    for _ in range(3):
        r = requests.post(f"{API}/auth/login", json={"email": email, "password": pwd}, timeout=15)
        if r.status_code == 200:
            return r.json()["access_token"]
        if r.status_code == 429:
            time.sleep(6)
            continue
        break
    raise AssertionError(f"login failed for {email}: {r.status_code} {r.text[:200]}")


@pytest.fixture(scope="module")
def admin_token():
    return _login(*ADMIN)


@pytest.fixture(scope="module")
def tech_token():
    return _login(*TECH)


@pytest.fixture(scope="module")
def viewer_token():
    return _login(*VIEWER)


def H(token):
    return {"Authorization": f"Bearer {token}"}


# ============ /api/reports/types ============
class TestReportTypes:
    def test_types_returns_4_known(self, tech_token):
        r = requests.get(f"{API}/reports/types", headers=H(tech_token), timeout=15)
        assert r.status_code == 200
        data = r.json()
        keys = {x["key"] for x in data}
        assert keys == {"plates", "events", "alerts", "equipment"}
        # equipment n'a pas de filtre date
        for x in data:
            if x["key"] == "equipment":
                assert x["date_filter"] is False
            else:
                assert x["date_filter"] is True

    def test_types_forbidden_for_viewer(self, viewer_token):
        r = requests.get(f"{API}/reports/types", headers=H(viewer_token), timeout=15)
        assert r.status_code == 403


# ============ Génération CSV/XLSX/PDF ============
@pytest.mark.parametrize("rtype", ["plates", "events", "alerts", "equipment"])
class TestGenerate:
    def test_csv(self, tech_token, rtype):
        r = requests.get(f"{API}/reports/{rtype}", headers=H(tech_token),
                         params={"format": "csv"}, timeout=30)
        assert r.status_code == 200, r.text[:200]
        assert "text/csv" in r.headers.get("content-type", "")
        assert len(r.content) > 0
        # 1ère ligne = en-têtes parsables
        reader = _csv.reader(io.StringIO(r.text))
        headers = next(reader)
        assert len(headers) > 0
        assert "attachment" in r.headers.get("content-disposition", "")

    def test_xlsx(self, tech_token, rtype):
        r = requests.get(f"{API}/reports/{rtype}", headers=H(tech_token),
                         params={"format": "xlsx"}, timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml.sheet" in r.headers.get("content-type", "")
        assert len(r.content) > 100
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.max_row >= 1  # au moins l'en-tête

    def test_pdf(self, tech_token, rtype):
        r = requests.get(f"{API}/reports/{rtype}", headers=H(tech_token),
                         params={"format": "pdf"}, timeout=45)
        assert r.status_code == 200
        assert "application/pdf" in r.headers.get("content-type", "")
        assert r.content.startswith(b"%PDF"), "Pas un PDF valide"
        assert len(r.content) > 500


# ============ Filtres ============
class TestFilters:
    def test_unknown_type_404(self, tech_token):
        r = requests.get(f"{API}/reports/unknown_xx", headers=H(tech_token),
                         params={"format": "csv"}, timeout=15)
        assert r.status_code == 404

    def test_invalid_format_400(self, tech_token):
        r = requests.get(f"{API}/reports/plates", headers=H(tech_token),
                         params={"format": "doc"}, timeout=15)
        assert r.status_code == 400

    def test_date_filter_reduces(self, tech_token):
        # Total sans filtre
        r_all = requests.get(f"{API}/reports/plates", headers=H(tech_token),
                             params={"format": "csv"}, timeout=30)
        all_lines = r_all.text.strip().splitlines()
        # Filtre vers le passé lointain (jeu vide attendu)
        r_old = requests.get(f"{API}/reports/plates", headers=H(tech_token),
                             params={"format": "csv", "date_from": "1990-01-01",
                                     "date_to": "1990-12-31"}, timeout=30)
        old_lines = r_old.text.strip().splitlines()
        assert r_old.status_code == 200
        # En-tête uniquement (filtre exclut tout)
        assert len(old_lines) == 1, f"attendu seulement en-têtes, got {len(old_lines)}"
        assert len(old_lines) < len(all_lines)

    def test_site_filter_reduces(self, tech_token):
        # Récupère un site_id
        sites = requests.get(f"{API}/sites", headers=H(tech_token), timeout=15).json()
        assert sites, "Aucun site"
        sid = sites[0]["id"]
        r_all = requests.get(f"{API}/reports/equipment", headers=H(tech_token),
                             params={"format": "csv"}, timeout=30)
        r_one = requests.get(f"{API}/reports/equipment", headers=H(tech_token),
                             params={"format": "csv", "site_id": sid}, timeout=30)
        assert r_all.status_code == 200 and r_one.status_code == 200
        all_n = len(r_all.text.strip().splitlines())
        one_n = len(r_one.text.strip().splitlines())
        # Le filtré doit être <= total
        assert one_n <= all_n


# ============ Sécurité ============
class TestSecurity:
    def test_viewer_forbidden_on_each_type(self, viewer_token):
        for rtype in ["plates", "events", "alerts", "equipment"]:
            r = requests.get(f"{API}/reports/{rtype}", headers=H(viewer_token),
                             params={"format": "csv"}, timeout=15)
            assert r.status_code == 403, f"{rtype} should be 403 for viewer"

    def test_unauth_401(self):
        r = requests.get(f"{API}/reports/plates", params={"format": "csv"}, timeout=15)
        assert r.status_code in (401, 403)


# ============ ANPR blacklist enrichie ============
class TestBlacklistAlertEnriched:
    def test_detect_blacklist_no_500(self, admin_token):
        # Avant
        a_before = requests.get(f"{API}/alerts", headers=H(admin_token),
                                params={"severity": "critical", "limit": 10}, timeout=15).json()
        r = requests.post(f"{API}/anpr/detect", headers=H(admin_token),
                          json={"plate": "ZZ-999-ZZ"}, timeout=15)
        assert r.status_code == 200, r.text[:200]
        body = r.json()
        assert body.get("list_status") == "black"
        assert body.get("blacklist_alert") is True
        # Laisse le tick background s'effectuer
        time.sleep(1.5)
        a_after = requests.get(f"{API}/alerts", headers=H(admin_token),
                               params={"severity": "critical", "limit": 10}, timeout=15).json()
        # Nouvelle alerte anpr_blacklist visible
        types_after = [a.get("type") for a in a_after]
        assert "anpr_blacklist" in types_after

    def test_detect_normal_plate_no_alert(self, admin_token):
        r = requests.post(f"{API}/anpr/detect", headers=H(admin_token),
                          json={"plate": "AA-111-AA"}, timeout=15)
        assert r.status_code == 200
        body = r.json()
        assert body.get("blacklist_alert") is False


# ============ Poll réseau ============
class TestNetworkPoll:
    def test_stats_coherent(self, tech_token):
        r = requests.get(f"{API}/network/stats", headers=H(tech_token), timeout=15)
        assert r.status_code == 200
        data = r.json()
        # clés cohérentes
        for k in ("total", "online", "offline"):
            assert k in data, f"clé manquante: {k}"
        assert data["total"] == data["online"] + data["offline"] + data.get("warning", 0)

    def test_manual_poll_works(self, tech_token):
        r = requests.post(f"{API}/network/poll", headers=H(tech_token), timeout=30)
        assert r.status_code == 200
        # accepte dict {polled: N} ou liste
        body = r.json()
        assert body is not None


# ============ Régression Recordings ZIP ============
class TestRecordingsExport:
    def test_export_zip(self, admin_token):
        cams = requests.get(f"{API}/cameras", headers=H(admin_token),
                            params={"status": "online"}, timeout=15).json()
        assert cams, "Aucune caméra"
        cid = cams[0]["id"]
        payload = {"camera_id": cid,
                   "start": "2025-01-15T08:00:00",
                   "end": "2025-01-15T08:05:00",
                   "format": "zip"}
        r = requests.post(f"{API}/recordings/export", headers=H(admin_token),
                          json=payload, timeout=60)
        assert r.status_code in (200, 201), r.text[:200]
        info = r.json()
        # télécharge
        eid = info.get("id") or info.get("export_id")
        assert eid
        dl = requests.get(f"{API}/recordings/exports/{eid}/download",
                          headers=H(admin_token), timeout=60)
        assert dl.status_code == 200
        assert len(dl.content) > 100
